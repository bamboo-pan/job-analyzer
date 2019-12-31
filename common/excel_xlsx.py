import datetime

import openpyxl
import os
import openpyxl.styles as style
from openpyxl.cell import cell

def read_excel_xlsx(file_name,sheetname="Sheet",start_row=1,start_column=1):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheetname]
    row_count: int = sheet.max_row
    column_count: int = sheet.max_column
    lists = []
    for i in range(start_row, row_count + 1):
        row_list=[]
        for j in range(start_column, column_count + 1):
            row_list.append(sheet.cell(row=i, column=j).value)
        lists.append(row_list)
    return lists

def get_excel_rows_cols_xls(file_name,sheetname="Sheet"):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    if not sheetname in wb.sheetnames:
        return None,None
    else:
        sheet = wb[sheetname]
        return sheet.max_row,sheet.max_column

def write_excel_xlsx_over_write(file_name,data,sheetname="Sheet",start_row=1,start_column=1):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    if not sheetname in wb.sheetnames:
        wb.create_sheet(sheetname)
    else:
        del wb[sheetname]
        wb.create_sheet(sheetname)
    sheet = wb[sheetname]
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet.cell(i+start_row,j+start_column).value=data[i][j]
    wb.save(file_name)


def write_excel_xlsx_append(file_name, data, sheetname="Sheet"):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    if not sheetname in wb.sheetnames:
        wb.create_sheet(sheetname)
    sheet = wb[sheetname]
    start_row=sheet.max_row
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet.cell(i + start_row+1, j+1).value = data[i][j]
    wb.save(file_name)

def generate_excel_file_name(key_word):
    return os.path.join(os.getcwd(),key_word+"_"+datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")+".xlsx")

def set_hyperlink(file_name,row,column,sheetname="Sheet",link=None):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheetname]
    if not link:
        sheet.cell(row=row, column=column).hyperlink=str(sheet.cell(row=row, column=column).value)
    else:
        sheet.cell(row=row, column=column).hyperlink=str(link)
    wb.save(file_name)

def set_hyperlink_range(file_name,ranges,sheetname="Sheet",link=None):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheetname]
    for i in range(ranges[0][0],ranges[0][1]):
        for j in range(ranges[1][0], ranges[1][1]):
            if not link:
                sheet.cell(row=i, column=j).hyperlink = str(sheet.cell(row=i, column=j).value)
            else:
                sheet.cell(row=i, column=j).hyperlink = str(link)
    wb.save(file_name)


def set_hyperlink_column(file_name,column,sheetname="Sheet",link=None):
    check_file_exist(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheetname]
    for i in range(1,sheet.max_row+1):
        if not link:
            sheet.cell(row=i, column=column).hyperlink = str(sheet.cell(row=i, column=column).value)
        else:
            sheet.cell(row=i, column=column).hyperlink = str(link)
    wb.save(file_name)


def set_columon_width_auto(file_name,sheetname="Sheet",start_row=1,start_column=1):
    try:
        if not os.path.exists(file_name):
            print("not found file")
            return None
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheetname]
        row_count: int = sheet.max_row
        column_count: int = sheet.max_column

        """
        获取所有单元格value宽度
        """
        lenth_list = []
        for i in range(start_row, row_count + 1):
            lenth_row=[]
            for j in range(start_column, column_count + 1):
                if sheet.cell(row=i, column=j).value:
                    lenth_row.append(len(str(sheet.cell(row=i, column=j).value)))
                else:
                    lenth_row.append(0)
            lenth_list.append(lenth_row)

        """
        由行list转为列list
        """
        new_list=[]
        for t in range(column_count):
            new_list.append([])
        for i in range(len(lenth_list)):
            for j in range(len(lenth_list[i])):
                new_list[j].append(lenth_list[i][j])

        """
        取列list最大值设置为列宽度
        """
        for i in range(1,len(new_list)+1):
           sheet.column_dimensions[cell.get_column_letter(i)].width=max(new_list[i-1])
        wb.save(file_name)

    except Exception as e :
        print(e)


def data_tag(file_name,sheetname="Sheet",start_row=1,start_column=1):
    if not os.path.exists(file_name):
        print("not found file")
        return None
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheetname]
    ft = style.Font(color=style.colors.RED)
    for i in range(start_row,sheet.max_row+1,2):
        for j in range(start_column,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value==sheet.cell(row=i+1, column=j).value:
                print("match found")
            else:
                sheet.cell(row=i, column=j).font=ft
                sheet.cell(row=i+1, column=j).font = ft
                print("not match found")
    wb.save(file_name)

def check_file_exist(file_name):
    if not os.path.exists(file_name):
        wb=openpyxl.Workbook()
        wb.save(file_name)



if __name__=="__main__":

    data=[[1,2,3,4],[2,2,2,2]]
    write_excel_xlsx_append("2.xlsx",data)
    print()



